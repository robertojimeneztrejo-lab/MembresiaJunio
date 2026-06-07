import streamlit as st
import google.generativeai as genai
import json
import re
import io
import requests
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="ARIA Membresías",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;500;600;700;800&family=DM+Sans:wght@300;400;500&display=swap');

:root {
    --canary: #F5E642;
    --canary-dark: #D4C800;
    --ink: #0F0F0F;
    --ink-soft: #2A2A2A;
    --surface: #F9F8F2;
    --muted: #6B6860;
    --border: #D8D5C8;
}

html, body, [data-testid="stAppViewContainer"] {
    background-color: var(--surface) !important;
    font-family: 'DM Sans', sans-serif;
    color: var(--ink);
}

[data-testid="stSidebar"] {
    background-color: var(--ink) !important;
    border-right: 3px solid var(--canary);
}

[data-testid="stSidebar"] * { color: #F0EDDE !important; }

[data-testid="stSidebar"] .stMarkdown h2,
[data-testid="stSidebar"] .stMarkdown h3 {
    font-family: 'Syne', sans-serif !important;
    color: var(--canary) !important;
}

.stCheckbox label { font-size: 0.82rem !important; font-family: 'DM Sans', sans-serif !important; }
[data-testid="stSidebar"] .stCheckbox label { color: #D0CDBE !important; }
[data-testid="stSidebar"] .stCheckbox label p { color: #D0CDBE !important; }

.main-header {
    font-family: 'Syne', sans-serif;
    font-size: 2.8rem;
    font-weight: 800;
    color: var(--ink);
    line-height: 1.05;
    letter-spacing: -0.03em;
}
.main-header span { background: var(--canary); padding: 0 6px; display: inline-block; }
.subtitle { font-size: 1rem; color: var(--muted); font-weight: 300; margin-top: 0.4rem; margin-bottom: 1.5rem; }

.result-card {
    background: white;
    border: 1.5px solid var(--border);
    border-radius: 12px;
    padding: 1.25rem 1.5rem;
    margin-bottom: 1rem;
    position: relative;
    overflow: hidden;
}
.result-card::before {
    content: '';
    position: absolute;
    top: 0; left: 0;
    width: 4px; height: 100%;
    background: var(--canary);
}
.result-card:hover { border-color: var(--canary-dark); }

.card-title { font-family: 'Syne', sans-serif; font-size: 1.1rem; font-weight: 700; color: var(--ink); margin-bottom: 0.2rem; }
.card-url { font-size: 0.78rem; color: var(--muted); margin-bottom: 0.8rem; word-break: break-all; }
.card-url a { color: #1a73e8; text-decoration: none; }
.card-section-title { font-size: 0.7rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.08em; color: var(--muted); margin-bottom: 0.25rem; margin-top: 0.7rem; }
.card-text { font-size: 0.88rem; color: var(--ink-soft); line-height: 1.55; }

.badge { font-size: 0.72rem; font-weight: 500; padding: 3px 10px; border-radius: 20px; font-family: 'DM Sans', sans-serif; }
.badge-type-academica { background: #EBF8FF; color: #1A6B99; border: 1px solid #BEE3F8; }
.badge-type-profesional { background: #F0FFF4; color: #22543D; border: 1px solid #9AE6B4; }
.badge-type-comercial { background: #FFFBEB; color: #7B4F12; border: 1px solid #F6D860; }
.badge-type-comunidad { background: #FAF5FF; color: #553C9A; border: 1px solid #D6BCFA; }
.badge-region { background: #F7FAFC; color: #2D3748; border: 1px solid #CBD5E0; }
.badge-access { background: #F0FFF4; color: #276749; border: 1px solid #9AE6B4; }

.stars { font-size: 1rem; color: var(--canary-dark); letter-spacing: 1px; }
.stars-empty { color: #D8D5C8; }

.stButton > button {
    background: var(--canary) !important;
    color: var(--ink) !important;
    font-family: 'Syne', sans-serif !important;
    font-weight: 700 !important;
    letter-spacing: 0.04em !important;
    border: 2px solid var(--ink) !important;
    border-radius: 6px !important;
    padding: 0.5rem 1.5rem !important;
    transition: all 0.15s !important;
}
.stButton > button:hover { background: var(--ink) !important; color: var(--canary) !important; }

.status-bar {
    background: var(--ink);
    color: var(--canary);
    font-family: 'Syne', sans-serif;
    font-size: 0.78rem;
    font-weight: 600;
    padding: 0.5rem 1rem;
    border-radius: 6px;
    letter-spacing: 0.06em;
    margin-bottom: 1.2rem;
    display: inline-block;
}

.empty-state { text-align: center; padding: 3rem 1rem; color: var(--muted); font-size: 0.95rem; }
.empty-icon { font-size: 2.5rem; margin-bottom: 0.8rem; display: block; }

.verified-badge { display: inline-block; font-size: 0.68rem; padding: 2px 8px; border-radius: 4px; margin-left: 8px; vertical-align: middle; }
.verified-ok { background: #F0FFF4; color: #276749; border: 1px solid #9AE6B4; }
.verified-no { background: #FFF5F5; color: #9B2C2C; border: 1px solid #FEB2B2; }
.verified-unk { background: #F7FAFC; color: #4A5568; border: 1px solid #CBD5E0; }

[data-testid="stSidebar"] .stSelectbox label,
[data-testid="stSidebar"] .stNumberInput label { color: #D0CDBE !important; font-size: 0.82rem !important; }

.action-row { display: flex; gap: 0.75rem; margin-bottom: 1.5rem; align-items: center; }
</style>
""", unsafe_allow_html=True)

# ── API Key desde Secrets ─────────────────────────────────────────────────────
api_key = st.secrets.get("GEMINI_API_KEY", "")

# ── Google Sheets URLs ───────────────────────────────────────────────────────
SHEETS_CSV_URL    = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTmA6CzuX4HnRv_gfKHdZB4GcezxNGq0g5nUY7OkFyEbPeYfkSbMgeSg7O20WoqPs-YRVF0qVc3AdRA/pub?output=csv"
APPS_SCRIPT_URL   = "https://script.google.com/macros/s/AKfycby-AfeStXZD4QF06uwEzpY3ZAG-8DyDABTYNohhoSdcYdf8dwDgK48W_K-Ou3ugp2WRCw/exec"

@st.cache_data(ttl=60)
def load_existing_memberships():
    """Lee membresías ya obtenidas desde Google Sheets (refresca cada 60 seg)."""
    try:
        df = pd.read_csv(SHEETS_CSV_URL, header=None)
        nombres = df.iloc[:, 0].dropna().astype(str).str.strip().tolist()
        if nombres and nombres[0].lower() in ["nombre", "name", "membresía", "membresia", "plataforma"]:
            nombres = nombres[1:]
        return [n for n in nombres if n]
    except Exception:
        return []


def save_to_sheets(results):
    """Envía los nombres de los resultados al Apps Script para guardarlos en el Sheet."""
    try:
        nombres = [r.get("nombre", "").strip() for r in results if r.get("nombre", "").strip()]
        if not nombres:
            return False, "Sin nombres para guardar"
        resp = requests.post(
            APPS_SCRIPT_URL,
            json={"nombres": nombres},
            timeout=15
        )
        data = resp.json()
        if data.get("status") == "ok":
            return True, data.get("added", 0)
        return False, data.get("message", "Error desconocido")
    except Exception as ex:
        return False, str(ex)

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🔍 ARIA\n### Membresías Gratuitas")
    st.markdown("---")

    # Indicador de lista de exclusión
    excluidas_preview = load_existing_memberships()
    n_exc = len(excluidas_preview)
    if n_exc > 0:
        st.markdown(
            f'''<div style="background:#1a1a1a;border:1px solid #F5E642;border-radius:8px;padding:0.6rem 0.8rem;margin-bottom:0.5rem;">
            <div style="font-size:0.65rem;letter-spacing:0.1em;text-transform:uppercase;color:#F5E642;font-weight:700;margin-bottom:3px;">📋 Lista de exclusión</div>
            <div style="font-size:0.82rem;color:#D0CDBE;"><strong style="color:#F5E642;">{n_exc}</strong> membresías ya obtenidas serán excluidas automáticamente</div>
            </div>''',
            unsafe_allow_html=True
        )
    else:
        st.markdown(
            '<div style="background:#1a1a1a;border:1px solid #333;border-radius:8px;padding:0.6rem 0.8rem;margin-bottom:0.5rem;">' +
            '<div style="font-size:0.65rem;letter-spacing:0.1em;text-transform:uppercase;color:#666;font-weight:700;margin-bottom:3px;">📋 Lista de exclusión</div>' +
            '<div style="font-size:0.82rem;color:#888;">Sin datos o cargando...</div></div>',
            unsafe_allow_html=True
        )
    st.markdown("---")

    st.markdown("### 📊 Resultados")
    num_results = st.slider("Cantidad de resultados", min_value=3, max_value=20, value=8, label_visibility="visible")

    st.markdown("---")
    st.markdown("### 🌎 Región prioritaria")
    reg_norteamerica = st.checkbox("Norteamérica (EE.UU. / Canadá)", value=True)
    reg_europa       = st.checkbox("Europa Occidental, Central y Norte", value=True)
    reg_latam        = st.checkbox("América Latina y el Caribe", value=True)

    st.markdown("---")
    st.markdown("### 📚 Tipo de membresía")
    tipo_academica  = st.checkbox("Académica / Institucional", value=True)
    tipo_profesional= st.checkbox("Profesiones y Oficios", value=True)
    tipo_comercial  = st.checkbox("Comercial / Herramientas", value=True)
    tipo_comunidad  = st.checkbox("Comunidad y Redes", value=True)

    st.markdown("---")
    st.markdown("### 👤 Audiencia objetivo")
    aud_instituciones = st.checkbox("Instituciones / Universidades", value=True)
    aud_investigadores= st.checkbox("Investigadores / Académicos", value=True)
    aud_docentes      = st.checkbox("Docentes / Profesores", value=True)
    aud_posgrado      = st.checkbox("Estudiantes de Posgrado", value=True)
    aud_licenciatura  = st.checkbox("Estudiantes de Licenciatura", value=True)
    aud_directivos    = st.checkbox("Directivos / Administradores", value=True)
    aud_bibliotecarios= st.checkbox("Bibliotecarios / Gestores", value=True)

    st.markdown("---")
    st.markdown("### 🔓 Condición de gratuidad")
    cond_total  = st.checkbox("Gratuidad total", value=True)
    cond_edu    = st.checkbox("Student Tier (.edu / dominio univ.)", value=True)
    cond_grant  = st.checkbox("Grant-Based Free Tier", value=True)
    cond_sandbox= st.checkbox("Institutional Sandbox", value=True)
    cond_alumni = st.checkbox("Año de Gracia / Alumni Launchpad", value=True)
    cond_cert   = st.checkbox("Gratuidad por Certificación", value=True)
    cond_beta   = st.checkbox("Acceso Beta Tester", value=True)
    cond_prueba = st.checkbox("Prueba Institucional 12 meses", value=True)

    st.markdown("---")
    st.markdown("### 🔐 Método de acceso")
    acc_sso      = st.checkbox("SSO Institucional / IP Whitelist", value=True)
    acc_edu      = st.checkbox("Dominio .edu verificado", value=True)
    acc_invitacion= st.checkbox("Invitación / Aprobación previa", value=True)
    acc_personal = st.checkbox("Cuenta personal validada", value=True)


# ── Helpers ───────────────────────────────────────────────────────────────────
def build_filters_summary():
    regiones, tipos, audiencias, condiciones, accesos = [], [], [], [], []
    if reg_norteamerica: regiones.append("Norteamérica (EE.UU. y Canadá)")
    if reg_europa:       regiones.append("Europa Occidental, Central y Norte")
    if reg_latam:        regiones.append("América Latina y el Caribe")
    if tipo_academica:   tipos.append("Académica/Institucional")
    if tipo_profesional: tipos.append("Profesiones y Oficios")
    if tipo_comercial:   tipos.append("Comercial/Herramientas")
    if tipo_comunidad:   tipos.append("Comunidad y Redes")
    if aud_instituciones:  audiencias.append("Instituciones y Universidades")
    if aud_investigadores: audiencias.append("Investigadores y Académicos")
    if aud_docentes:       audiencias.append("Docentes y Profesores")
    if aud_posgrado:       audiencias.append("Estudiantes de Posgrado")
    if aud_licenciatura:   audiencias.append("Estudiantes de Licenciatura")
    if aud_directivos:     audiencias.append("Directivos y Administradores Universitarios")
    if aud_bibliotecarios: audiencias.append("Bibliotecarios y Gestores de Repositorios")
    if cond_total:   condiciones.append("Gratuidad total")
    if cond_edu:     condiciones.append("Student Tier (.edu o dominio universitario)")
    if cond_grant:   condiciones.append("Grant-Based Free Tier")
    if cond_sandbox: condiciones.append("Institutional Sandbox")
    if cond_alumni:  condiciones.append("Año de Gracia / Alumni Launchpad")
    if cond_cert:    condiciones.append("Gratuidad por certificación o cursos completados")
    if cond_beta:    condiciones.append("Acceso Beta Tester")
    if cond_prueba:  condiciones.append("Prueba institucional de 12 meses")
    if acc_sso:        accesos.append("SSO Institucional o IP Whitelisting")
    if acc_edu:        accesos.append("Dominio .edu verificado")
    if acc_invitacion: accesos.append("Invitación o aprobación previa")
    if acc_personal:   accesos.append("Cuenta personal con validación de perfil")
    return regiones, tipos, audiencias, condiciones, accesos


def build_prompt(topic, regiones, tipos, audiencias, condiciones, accesos, n, excluidas=None):
    return f"""Eres un agente especializado en inteligencia de recursos académicos y profesionales. Tu función es identificar, evaluar y catalogar plataformas web que ofrecen membresías, suscripciones o accesos institucionales COMPLETAMENTE GRATUITOS para el sector educativo y de investigación.

TEMA DE BÚSQUEDA: {topic if topic else "herramientas y recursos académicos generales"}

OBJETIVO: Encuentra exactamente {n} páginas web que ofrezcan membresías académicas o institucionales GRATUITAS (sin costo alguno, periodo mínimo 12 meses o 1 año) para el tema indicado.

REGIONES PRIORITARIAS (busca SOLO en estas):
{chr(10).join(f"- {r}" for r in regiones) if regiones else "- Sin filtro regional"}

TIPOS DE MEMBRESÍA A BUSCAR:
{chr(10).join(f"- {t}" for t in tipos) if tipos else "- Todos los tipos"}

AUDIENCIAS OBJETIVO:
{chr(10).join(f"- {a}" for a in audiencias) if audiencias else "- Todas las audiencias"}

CONDICIONES DE GRATUIDAD ACEPTABLES (al menos una debe cumplirse):
{chr(10).join(f"- {c}" for c in condiciones) if condiciones else "- Cualquier condición de gratuidad"}

MÉTODOS DE ACCESO ACEPTABLES:
{chr(10).join(f"- {a}" for a in accesos) if accesos else "- Cualquier método de acceso"}

REGLAS OBLIGATORIAS:
1. Solo membresías COMPLETAMENTE GRATUITAS. Excluye cualquier opción de pago.
2. El período de acceso gratuito debe ser de 12 meses o 1 año como mínimo.
3. Excluye plataformas de: Asia Oriental/Pacífico, Asia del Sur/Central, África, Medio Oriente.
4. Excluye organismos multilaterales: ONU, UNESCO, FMI, BM, OCDE, OMS, OEA y equivalentes.
5. Prioriza plataformas con poca visibilidad sobre las ampliamente conocidas.
6. Nunca inventes URLs. Si no puedes verificar un dato, usa "Sin verificar".
7. No incluyas contenido detrás de login previo que no sea descubrible públicamente.
8. Ordena los resultados de mayor a menor puntuación (5 a 1).
9. NUNCA repitas las siguientes membresías que ya han sido obtenidas previamente (lista de exclusión):
{{EXCLUSION_BLOCK}}

FORMATO DE RESPUESTA: Responde ÚNICAMENTE con un array JSON válido, sin texto adicional, sin bloques de código markdown. El array debe tener exactamente {n} objetos:

[
  {{
    "nombre": "Nombre oficial de la plataforma",
    "url": "https://url-directa-a-pagina-de-membresia.com",
    "para_quien_es_gratis": "Descripción exacta del perfil que califica",
    "condicion_gratuidad": "Tipo específico de gratuidad que aplica",
    "metodo_acceso": "Método de acceso requerido",
    "beneficios": ["Beneficio 1", "Beneficio 2", "Beneficio 3", "Beneficio 4"],
    "tipo_membresia": "Académica|Profesional|Comercial|Comunidad",
    "region": "Norteamérica|Europa|América Latina|Global",
    "puntuacion": 4,
    "url_verificada": true,
    "duracion": "12 meses / 1 año / Indefinida mientras seas estudiante"
  }}
]""".replace("{{EXCLUSION_BLOCK}}", ("\n".join(f"- {e}" for e in excluidas) if excluidas else "ninguna por ahora"))


def clean_json_string(raw):
    raw = re.sub(r"```json", "", raw)
    raw = re.sub(r"```", "", raw).strip()
    raw = re.sub(r",\s*([}\]])", r"\1", raw)
    raw = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", raw)
    m = re.search(r"\[.*\]", raw, re.DOTALL)
    if m:
        raw = m.group()
    return raw


def parse_json_results(raw):
    """Intenta parsear el JSON; si está truncado, rescata los objetos completos."""
    raw = clean_json_string(raw)

    # Intento 1: parseo completo
    try:
        results = json.loads(raw)
        return results if isinstance(results, list) else [results]
    except json.JSONDecodeError:
        pass

    # Intento 2: cerrar JSON truncado — añadir }] al final y reintentar
    for suffix in ["}]", "}]}]", "]"]:
        try:
            results = json.loads(raw + suffix)
            return results if isinstance(results, list) else [results]
        except json.JSONDecodeError:
            pass

    # Intento 3: extraer objetos completos uno a uno
    # Busca llaves balanceadas para no cortar campos a medias
    results = []
    depth = 0
    start = None
    for i, ch in enumerate(raw):
        if ch == "{":
            if depth == 0:
                start = i
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0 and start is not None:
                fragment = raw[start:i+1]
                fragment = clean_json_string(fragment)
                try:
                    results.append(json.loads(fragment))
                except json.JSONDecodeError:
                    pass
                start = None
    if results:
        return results

    raise ValueError(
        f"No se pudo parsear la respuesta de Gemini.\n"
        f"Intenta reducir el número de resultados o vuelve a buscar.\n\n"
        f"Fragmento recibido:\n{raw[:300]}"
    )


def run_search(topic, regiones, tipos, audiencias, condiciones, accesos, n):
    excluidas = load_existing_memberships()
    prompt = build_prompt(topic, regiones, tipos, audiencias, condiciones, accesos, n, excluidas)
    genai.configure(api_key=api_key)
    model = genai.GenerativeModel("gemini-2.5-flash")

    # Calcular tokens necesarios: ~400 tokens por resultado + margen
    needed_tokens = max(8192, n * 500 + 2000)

    response = model.generate_content(
        prompt,
        generation_config=genai.types.GenerationConfig(
            temperature=0.3,
            max_output_tokens=needed_tokens
        )
    )
    raw = response.text.strip()

    # Detectar respuesta truncada por finish_reason
    finish = None
    try:
        finish = response.candidates[0].finish_reason.name
    except Exception:
        pass

    results = parse_json_results(raw)

    # Si vino truncada y recuperamos menos resultados de los pedidos, avisar en sesión
    if finish == "MAX_TOKENS" and len(results) < n:
        st.session_state["truncated_warning"] = (
            f"⚠️ Gemini truncó la respuesta ({len(results)} de {n} resultados recuperados). "
            "Intenta pedir menos resultados o pulsa 🔄 Nuevos resultados."
        )
    else:
        st.session_state.pop("truncated_warning", None)

    return results


def render_stars(score):
    return f'<span class="stars">{"★" * score}</span><span class="stars stars-empty">{"☆" * (5 - score)}</span>'


def render_type_badge(tipo):
    t = (tipo or "").lower()
    if "académic" in t or "institucional" in t:
        return f'<span class="badge badge-type-academica">{tipo}</span>'
    elif "profesion" in t or "oficio" in t:
        return f'<span class="badge badge-type-profesional">{tipo}</span>'
    elif "comercial" in t or "herramienta" in t:
        return f'<span class="badge badge-type-comercial">{tipo}</span>'
    else:
        return f'<span class="badge badge-type-comunidad">{tipo}</span>'


def render_results(results):
    for i, r in enumerate(results):
        score    = r.get("puntuacion", 3)
        verified = r.get("url_verificada", None)
        if verified is True:
            ver_html = '<span class="verified-badge verified-ok">✓ URL verificada</span>'
        elif verified is False:
            ver_html = '<span class="verified-badge verified-no">⚠ Sin verificar</span>'
        else:
            ver_html = '<span class="verified-badge verified-unk">? Por verificar</span>'

        url       = r.get("url", "#")
        nombre    = r.get("nombre", "Sin nombre")
        bene_html = "".join(f"<li>{b}</li>" for b in r.get("beneficios", []))

        st.markdown(f"""
        <div class="result-card">
            <div class="card-title">#{i+1} — {nombre} {ver_html}</div>
            <div class="card-url"><a href="{url}" target="_blank">{url}</a></div>
            <div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap;margin-bottom:0.6rem;">
                {render_stars(score)}
                {render_type_badge(r.get("tipo_membresia",""))}
                <span class="badge badge-region">{r.get("region","")}</span>
                <span class="badge badge-access">🕐 {r.get("duracion","12 meses")}</span>
            </div>
            <div class="card-section-title">Para quién es gratis</div>
            <div class="card-text">{r.get("para_quien_es_gratis","")}</div>
            <div class="card-section-title">Condición de gratuidad</div>
            <div class="card-text">{r.get("condicion_gratuidad","")}</div>
            <div class="card-section-title">Método de acceso</div>
            <div class="card-text">{r.get("metodo_acceso","")}</div>
            <div class="card-section-title">Beneficios principales</div>
            <div class="card-text"><ul style="margin:0;padding-left:1.2rem;">{bene_html}</ul></div>
        </div>
        """, unsafe_allow_html=True)


def build_excel(results, topic):
    wb = Workbook()
    ws = wb.active
    ws.title = "Membresías Gratuitas"

    # Paleta
    COLOR_HEADER_BG = "0F0F0F"
    COLOR_HEADER_FG = "F5E642"
    COLOR_SUBHEADER  = "F5E642"
    COLOR_SUBHEADER_FG = "0F0F0F"
    COLOR_ROW_ALT   = "F9F8F2"
    COLOR_BORDER    = "D8D5C8"
    COLOR_SCORE_5   = "1DB954"
    COLOR_SCORE_4   = "6BCB77"
    COLOR_SCORE_3   = "F5E642"
    COLOR_SCORE_2   = "FFB347"
    COLOR_SCORE_1   = "E53E3E"

    thin = Side(style="thin", color=COLOR_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Título
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = f"ARIA Membresías Gratuitas — {topic or 'Recursos Académicos'}"
    title_cell.font = Font(name="Arial", bold=True, size=14, color=COLOR_HEADER_FG)
    title_cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Fecha
    ws.merge_cells("A2:K2")
    date_cell = ws["A2"]
    date_cell.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Total resultados: {len(results)}"
    date_cell.font = Font(name="Arial", size=10, color="6B6860")
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Encabezados
    headers = ["#", "Nombre", "URL", "Para quién es gratis", "Condición de gratuidad",
               "Método de acceso", "Beneficios", "Tipo", "Región", "Duración", "Puntuación"]
    col_widths = [4, 28, 40, 35, 30, 28, 55, 18, 16, 22, 12]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color=COLOR_SUBHEADER_FG)
        cell.fill = PatternFill("solid", fgColor=COLOR_SUBHEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[3].height = 22

    # Datos
    score_colors = {5: COLOR_SCORE_5, 4: COLOR_SCORE_4, 3: COLOR_SCORE_3, 2: COLOR_SCORE_2, 1: COLOR_SCORE_1}

    for row_idx, r in enumerate(results, start=4):
        alt = row_idx % 2 == 0
        row_fill = PatternFill("solid", fgColor=COLOR_ROW_ALT) if alt else PatternFill("solid", fgColor="FFFFFF")

        beneficios_txt = "\n".join(f"• {b}" for b in r.get("beneficios", []))
        score = r.get("puntuacion", 3)
        stars = "★" * score + "☆" * (5 - score)

        values = [
            row_idx - 3,
            r.get("nombre", ""),
            r.get("url", ""),
            r.get("para_quien_es_gratis", ""),
            r.get("condicion_gratuidad", ""),
            r.get("metodo_acceso", ""),
            beneficios_txt,
            r.get("tipo_membresia", ""),
            r.get("region", ""),
            r.get("duracion", ""),
            stars,
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True,
                                       horizontal="center" if col_idx in [1, 8, 9, 10, 11] else "left")

            # Colorear puntuación
            if col_idx == 11:
                sc = score_colors.get(score, "888888")
                cell.font = Font(name="Arial", size=10, bold=True, color=sc)
                cell.fill = row_fill

        ws.row_dimensions[row_idx].height = max(60, 15 * len(r.get("beneficios", [])))

    # Freeze panes
    ws.freeze_panes = "A4"

    # Hoja de metadatos
    ws2 = wb.create_sheet("Filtros aplicados")
    ws2["A1"] = "Parámetro"
    ws2["B1"] = "Valor"
    for cell in [ws2["A1"], ws2["B1"]]:
        cell.font = Font(name="Arial", bold=True, size=10, color=COLOR_SUBHEADER_FG)
        cell.fill = PatternFill("solid", fgColor=COLOR_SUBHEADER)
        cell.border = border
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 60

    meta = [
        ("Tema de búsqueda", topic or "Recursos académicos generales"),
        ("Fecha de generación", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Total resultados", len(results)),
        ("Regiones", ", ".join([r for r in ["Norteamérica" if reg_norteamerica else "", "Europa" if reg_europa else "", "América Latina" if reg_latam else ""] if r])),
        ("Tipos de membresía", ", ".join([t for t, v in [("Académica", tipo_academica), ("Profesional", tipo_profesional), ("Comercial", tipo_comercial), ("Comunidad", tipo_comunidad)] if v])),
        ("Condiciones de gratuidad", ", ".join([c for c, v in [("Gratuidad total", cond_total), ("Student Tier", cond_edu), ("Grant-Based", cond_grant), ("Sandbox", cond_sandbox), ("Alumni Launchpad", cond_alumni), ("Certificación", cond_cert), ("Beta Tester", cond_beta), ("Prueba 12m", cond_prueba)] if v])),
    ]
    for r_idx, (k, v) in enumerate(meta, start=2):
        ws2.cell(row=r_idx, column=1, value=k).font = Font(name="Arial", size=9, bold=True)
        ws2.cell(row=r_idx, column=2, value=str(v)).font = Font(name="Arial", size=9)
        ws2.cell(row=r_idx, column=1).border = border
        ws2.cell(row=r_idx, column=2).border = border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Main UI ───────────────────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
    ARIA<br><span>Membresías</span><br>Gratuitas
</div>
<div class="subtitle">Inteligencia de recursos académicos y profesionales — acceso sin costo</div>
""", unsafe_allow_html=True)

topic_col, btn_col = st.columns([4, 1])
with topic_col:
    topic = st.text_input(
        "Tema o dominio de búsqueda",
        placeholder="Ej: software de análisis estadístico, gestión de referencias, diseño curricular...",
    )
with btn_col:
    st.markdown("<br>", unsafe_allow_html=True)
    buscar = st.button("🔍 Buscar", use_container_width=True)

# ── Session state ─────────────────────────────────────────────────────────────
if "results" not in st.session_state:
    st.session_state.results = None
if "last_topic" not in st.session_state:
    st.session_state.last_topic = ""

def do_search():
    if not api_key:
        st.error("Configura GEMINI_API_KEY en los Secrets de Streamlit Cloud.")
        return
    regiones, tipos, audiencias, condiciones, accesos = build_filters_summary()
    if not regiones:
        st.warning("Selecciona al menos una región en el panel lateral.")
        return
    if not condiciones:
        st.warning("Selecciona al menos una condición de gratuidad.")
        return
    with st.spinner("Buscando membresías gratuitas..."):
        try:
            results = run_search(topic, regiones, tipos, audiencias, condiciones, accesos, num_results)
            st.session_state.results = results
            st.session_state.last_topic = topic
        except Exception as e:
            st.error(f"Error: {str(e)}")

if buscar:
    do_search()

# ── Mostrar resultados ────────────────────────────────────────────────────────
if st.session_state.results:
    results = st.session_state.results

    # Barra de acciones
    col_status, col_regen, col_excel, col_save = st.columns([2.5, 1.2, 1.2, 1.2])
    with col_status:
        st.markdown(
            f'<div class="status-bar">✦ {len(results)} membresías gratuitas encontradas</div>',
            unsafe_allow_html=True
        )
    with col_regen:
        if st.button("🔄 Nuevos resultados", use_container_width=True):
            do_search()
            st.rerun()
    with col_excel:
        excel_buf = build_excel(results, st.session_state.last_topic)
        fname = f"ARIA_Membresias_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.download_button(
            label="📥 Descargar Excel",
            data=excel_buf,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    with col_save:
        if st.button("💾 Guardar en Sheet", use_container_width=True):
            with st.spinner("Guardando en Google Sheets..."):
                ok, info = save_to_sheets(results)
            if ok:
                added = info if isinstance(info, int) else 0
                if added > 0:
                    st.success(f"✓ {added} nuevas membresías guardadas")
                else:
                    st.info("Todas ya estaban en el Sheet")
                st.cache_data.clear()
                st.rerun()
            else:
                st.error(f"Error al guardar: {info}")

    # Mostrar advertencia de truncado si aplica
    if st.session_state.get("truncated_warning"):
        st.warning(st.session_state["truncated_warning"])

    render_results(results)

else:
    st.markdown("""
    <div class="empty-state">
        <span class="empty-icon">◈</span>
        Configura los filtros en el panel lateral y escribe un tema para comenzar.<br>
        <span style="font-size:0.82rem; color:#9E9B90;">Los filtros siempre están visibles — ajústalos antes de cada búsqueda.</span>
    </div>
    """, unsafe_allow_html=True)
